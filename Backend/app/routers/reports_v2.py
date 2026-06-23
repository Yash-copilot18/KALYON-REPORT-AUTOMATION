# app/routers/reports_v2.py

from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import Optional
import io
import csv
import logging
from datetime import datetime

from app.database.session import get_db
from app.schemas.reports_schema import ReportDataRequest
from app.services.reports_service import ReportsService

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/reports-v2", tags=["Reports v2"])

# ── Unit map ──────────────────────────────────────────────────────────────────
UNIT_MAP = {
    "DC_VOLTAGE": "V", "DC_CURRENT": "A", "DC_POWER": "kW",
    "ACTIVE_POWER": "kW", "REACTIVE_POWER": "kVAR", "APPARENT_POWER": "kVA",
    "DAILY_ENERGY": "kWh", "MONTHLY_ENERGY": "kWh", "YEARLY_ENERGY": "kWh",
    "LIFETIME_ENERGY": "kWh", "DAILY_REACT_ENERGY": "kVARh",
    "GRID_CURRENT_PHASE1": "A", "GRID_CURRENT_PHASE2": "A", "GRID_CURRENT_PHASE3": "A",
    "GRID_LINE_VOLT_UV": "V", "GRID_LINE_VOLT_VW": "V", "GRID_LINE_VOLT_WU": "V",
    "GRID_FREQUENCY_MEASURED": "Hz", "GRID_PF_MEASURED": "",
    "GRID_ACTIVE_POWER_MEASURED": "kW", "GRID_REACTIVE_POWER_MEASURED": "kVAR",
    "GRID_VOLTAGE_L_L_MEASURED": "V", "INVERTER_TOTAL_ACTIVE_POWER": "kW",
    "INVERTER_TOTAL_REACTIVE_POWER": "kVAR", "PLANT_DAILY_PRODUCTION": "kWh",
    "PLANT_MONTHLY_PRODUCTION": "kWh", "PLANT_YEARLY_PRODUCTION": "kWh",
    "PLANT_YEARLY_PRODUCTION": "kWh", "PLANT_LIFETIME_PRODUCTION": "kWh",
    "AVG_GHI_IRRADIATION": "W/m2", "AVG_GTI_IRRADIATION": "W/m2",
    "TOTAL_IRRADIANCE": "W/m2", "AVG_AIR_TEMP": "C",
    "AVG_WIND_SPEED": "m/s", "AVG_RELATIVE_HUMIDITY": "%",
    "AVG_IR_SOILING_RATIO1": "%", "ALL_WMS_AVG_MODULE_TEMP": "C",
    "AVG_AIR_PRESSURE": "hPa", "PRIMEPACK_IGBT_HEATSINK_TEMP": "C",
    "MV_TRAFO_TEMP": "C", "DAILY_OPERATING_TIME": "min",
    "INVERTER_RUNNING": "", "STRING_CURRENT1": "A", "STRING_CURRENT2": "A",
    "STRING_CURRENT3": "A", "STRING_CURRENT4": "A", "STRING_CURRENT5": "A",
    "DAILY_RUN_MIN": "min", "DOWN_TIME_MIN": "min", "OK_TIME_MIN": "min",
    "GRID_CURRENT_Ia": "A", "GRID_CURRENT_Ib": "A", "GRID_CURRENT_Ic": "A",
    "ACTIVE_POWER_SET_POINT": "kW", "VOLTAGE_SET_POINT": "V",
    "VAR_SET_POINT": "kVAR", "POWER_FACTOR_SET_POINT": "",
    "SCB1_DC_POWER": "kW", "SCB1_DC_VOLTAGE": "V", "SCB1_INTERNAL_TEMP": "C",
    "SCB1_TOTAL_CURRENT": "A", "SCB2_DC_POWER": "kW", "SCB2_DC_VOLTAGE": "V",
    "DurCol": "s",
}


def fmt_dmy(val):
    """Format any value to DD/MM/YYYY HH:MM:SS."""
    if not val:
        return ""
    if isinstance(val, str):
        val = val.replace("T", " ").replace("Z", "").strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
            try:
                val = datetime.strptime(val[:len(fmt)], fmt)
                break
            except Exception:
                continue
        if isinstance(val, str):
            return val
    if hasattr(val, "strftime"):
        return val.strftime("%d/%m/%Y %H:%M:%S")
    return str(val)


def fmt_val(val):
    """Format numeric to 3 decimal places."""
    if val is None:
        return ""
    if isinstance(val, float):
        return round(val, 3)
    if isinstance(val, int):
        return val
    return val


def fmt_val_str(val):
    """Format numeric to 3dp string for CSV/display."""
    if val is None:
        return ""
    if isinstance(val, float):
        return f"{val:.3f}"
    if isinstance(val, int):
        return str(val)
    return str(val)


def col_header(col: str) -> str:
    """Column name with unit."""
    if col == "timestamp":
        return "Timestamp (DD/MM/YYYY HH:MM:SS)"
    label = col.replace("_", " ").title()
    unit  = UNIT_MAP.get(col, "")
    return f"{label} ({unit})" if unit else label


def today_dmy():
    return datetime.now().strftime("%d-%m-%Y")


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.get("/equipment-types")
def get_equipment_types(db: Session = Depends(get_db)):
    return ReportsService.get_equipment_types(db)


@router.get("/equipment-list")
def get_equipment_list(type: str = Query(...), db: Session = Depends(get_db)):
    return ReportsService.get_equipment_list(db, type)


@router.get("/tags")
def get_tags(
    equipment_type: str = Query(...),
    equipment_id: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
):
    return ReportsService.get_tags(db, equipment_type, equipment_id)


@router.post("/data")
def get_report_data(req: ReportDataRequest, db: Session = Depends(get_db)):
    return ReportsService.get_report_data(db, req)


# ── Export CSV ────────────────────────────────────────────────────────────────
@router.post("/export/csv")
def export_csv(req: ReportDataRequest, db: Session = Depends(get_db)):
    req.page = 1
    req.page_size = 10000
    result = ReportsService.get_report_data(db, req)
    rows   = result.get("rows", [])
    cols   = result.get("columns", [])

    output = io.StringIO()
    output.write('\ufeff')  # UTF-8 BOM
    writer = csv.writer(output, delimiter=',', quoting=csv.QUOTE_MINIMAL)

    # ── Title block ───────────────────────────────────────────────────────
    writer.writerow(["GE Solar Monitoring & Report Automation"])
    writer.writerow([])
    writer.writerow(["Equipment Type",    req.equipment_type])
    writer.writerow(["Equipment ID",      req.equipment_id])
    writer.writerow(["From Date",         fmt_dmy(req.from_datetime)])
    writer.writerow(["To Date",           fmt_dmy(req.to_datetime)])
    writer.writerow(["Time Interval",     req.interval])
    writer.writerow(["Aggregation",       req.agg_function.upper()])
    writer.writerow(["Total Records",     result.get("total_records", 0)])
    writer.writerow(["Generated",         datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    writer.writerow([])

    # ── Header row ────────────────────────────────────────────────────────
    writer.writerow([col_header(c) for c in cols])

    # ── Data rows ─────────────────────────────────────────────────────────
    for row in rows:
        data_row = []
        for col in cols:
            val = row.get(col)
            if col == "timestamp":
                data_row.append(fmt_dmy(val))
            else:
                data_row.append(fmt_val_str(val))
        writer.writerow(data_row)

    output.seek(0)
    filename = f"{req.equipment_id}_Report_{today_dmy()}.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )


# ── Export Excel ──────────────────────────────────────────────────────────────
@router.post("/export/excel")
def export_excel(req: ReportDataRequest, db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, detail="Run: pip install openpyxl")

    req.page = 1
    req.page_size = 10000
    result = ReportsService.get_report_data(db, req)
    rows   = result.get("rows", [])
    cols   = result.get("columns", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Report Data"
    ws.sheet_view.showGridLines = False

    # ── Color constants ───────────────────────────────────────────────────
    NAVY    = "0F1524"
    SURFACE = "141928"
    CARD    = "1A2035"
    ACCENT  = "00D4AA"
    TEXT1   = "E8ECF4"
    TEXT2   = "A0AEC0"
    TEXT3   = "6B7A99"
    BORDER  = "2A3350"
    EVEN    = "1A2035"
    ODD     = "141928"
    BLUE    = "0066CC"

    def fill(c): return PatternFill("solid", fgColor=c)
    def font(c, sz=10, bold=False, italic=False):
        return Font(color=c, size=sz, bold=bold, italic=italic, name="Calibri")
    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    thin_side   = Side(style="thin",   color=BORDER)
    medium_side = Side(style="medium", color=ACCENT)
    thin_border = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=thin_side)
    hdr_border  = Border(left=thin_side, right=thin_side,
                         top=thin_side, bottom=medium_side)

    n = len(cols)

    def merge_row(r, val, fnt, aln, fll, h=20):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
        c = ws.cell(row=r, column=1, value=val)
        c.font = fnt; c.alignment = aln; c.fill = fll
        ws.row_dimensions[r].height = h

    # ── Row 1 — Company ───────────────────────────────────────────────────
    merge_row(1,
        "GE Solar Monitoring & Report Automation",
        font(ACCENT, 14, bold=True),
        align("center"),
        fill(NAVY), h=32)

    # ── Row 2 — Report title ──────────────────────────────────────────────
    merge_row(2,
        f"{req.equipment_type} — {req.equipment_id} — Data Report",
        font(TEXT1, 11, bold=True),
        align("center"),
        fill(SURFACE), h=22)

    # ── Rows 3–8 — Metadata ───────────────────────────────────────────────
    meta = [
        ("Equipment Type",    req.equipment_type),
        ("Equipment ID",      req.equipment_id),
        ("From Date",         fmt_dmy(req.from_datetime)),
        ("To Date",           fmt_dmy(req.to_datetime)),
        ("Time Interval",     req.interval),
        ("Aggregation",       req.agg_function.upper()),
        ("Total Records",     f"{result.get('total_records', 0):,}"),
        ("Generated",         datetime.now().strftime("%d/%m/%Y %H:%M:%S")),
    ]
    half = n // 2 if n > 1 else 1
    for i, (k, v) in enumerate(meta):
        r = 3 + (i // 2)
        col_start = 1 if i % 2 == 0 else half + 1
        col_end   = half if i % 2 == 0 else n

        ws.merge_cells(start_row=r, start_column=col_start,
                       end_row=r, end_column=col_end)
        c = ws.cell(row=r, column=col_start,
                    value=f"  {k}:  {v}")
        c.font      = font(TEXT2, 9)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill      = fill(CARD)
        ws.row_dimensions[r].height = 18

    META_END = 3 + (len(meta) - 1) // 2

    # ── Spacer ────────────────────────────────────────────────────────────
    SPACER_ROW = META_END + 1
    ws.row_dimensions[SPACER_ROW].height = 6

    # ── Header row ────────────────────────────────────────────────────────
    HDR_ROW = SPACER_ROW + 1
    for ci, col in enumerate(cols, start=1):
        c = ws.cell(row=HDR_ROW, column=ci, value=col_header(col))
        c.font      = font(ACCENT, 10, bold=True)
        c.fill      = fill(NAVY)
        c.alignment = align("center", "center", wrap=True)
        c.border    = hdr_border
    ws.row_dimensions[HDR_ROW].height = 38

    # ── Data rows ─────────────────────────────────────────────────────────
    for ri, row in enumerate(rows, start=HDR_ROW + 1):
        bg = EVEN if (ri - HDR_ROW) % 2 == 0 else ODD
        ws.row_dimensions[ri].height = 16

        for ci, col in enumerate(cols, start=1):
            val = row.get(col)
            c   = ws.cell(row=ri, column=ci)
            c.fill   = fill(bg)
            c.border = thin_border

            if col == "timestamp":
                c.value        = fmt_dmy(val) if val else "—"
                c.font         = font(TEXT3, 9)
                c.alignment    = align("center")
            elif val is None:
                c.value        = "—"
                c.font         = font(TEXT3, 9)
                c.alignment    = align("center")
            elif isinstance(val, float):
                c.value        = round(val, 3)
                c.font         = font(ACCENT, 9)
                c.alignment    = align("right")
                c.number_format = '#,##0.000'
            elif isinstance(val, int):
                c.value        = val
                c.font         = font(ACCENT, 9)
                c.alignment    = align("right")
                c.number_format = '#,##0'
            else:
                c.value        = str(val)
                c.font         = font(TEXT2, 9)
                c.alignment    = align("left")

    # ── Column widths ─────────────────────────────────────────────────────
    for ci, col in enumerate(cols, start=1):
        letter = get_column_letter(ci)
        hlen   = len(col_header(col))
        if col == "timestamp":
            ws.column_dimensions[letter].width = 24
        elif hlen > 30:
            ws.column_dimensions[letter].width = 22
        elif hlen > 20:
            ws.column_dimensions[letter].width = 18
        else:
            ws.column_dimensions[letter].width = 15

    # ── Freeze + autofilter ───────────────────────────────────────────────
    ws.freeze_panes = f"A{HDR_ROW + 1}"
    ws.auto_filter.ref = (
        f"A{HDR_ROW}:{get_column_letter(n)}{HDR_ROW + len(rows)}"
    )
    ws.sheet_properties.tabColor = ACCENT

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"{req.equipment_id}_Report_{today_dmy()}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )


# ── Export PDF ────────────────────────────────────────────────────────────────
@router.post("/export/pdf")
def export_pdf(req: ReportDataRequest, db: Session = Depends(get_db)):
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            SimpleDocTemplate, Table, TableStyle,
            Paragraph, Spacer, HRFlowable
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise HTTPException(500, detail="Run: pip install reportlab")

    req.page = 1
    req.page_size = 2000
    result = ReportsService.get_report_data(db, req)
    rows   = result.get("rows", [])
    cols   = result.get("columns", [])

    output    = io.BytesIO()
    page_size = landscape(A4)
    pw, ph    = page_size

    # Colors
    C_NAVY   = colors.HexColor("#0F1524")
    C_CARD   = colors.HexColor("#141928")
    C_ACCENT = colors.HexColor("#00D4AA")
    C_TEXT1  = colors.HexColor("#E8ECF4")
    C_TEXT2  = colors.HexColor("#A0AEC0")
    C_TEXT3  = colors.HexColor("#6B7A99")
    C_BORDER = colors.HexColor("#2A3350")
    C_EVEN   = colors.HexColor("#1A2035")
    C_ODD    = colors.HexColor("#141928")

    gen_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    page_num = [0]

    def on_page(canvas, doc):
        page_num[0] += 1
        canvas.saveState()

        # Header band
        canvas.setFillColor(C_NAVY)
        canvas.rect(0, ph - 1.3*cm, pw, 1.3*cm, fill=1, stroke=0)

        canvas.setFillColor(C_ACCENT)
        canvas.setFont("Helvetica-Bold", 11)
        canvas.drawString(1*cm, ph - 0.9*cm,
                          "GE Solar Monitoring & Report Automation")

        canvas.setFillColor(C_TEXT3)
        canvas.setFont("Helvetica", 8)
        canvas.drawRightString(pw - 1*cm, ph - 0.9*cm,
                               f"Page {page_num[0]}   |   Generated: {gen_time}")

        # Accent underline
        canvas.setStrokeColor(C_ACCENT)
        canvas.setLineWidth(1.5)
        canvas.line(0, ph - 1.3*cm, pw, ph - 1.3*cm)

        # Footer band
        canvas.setFillColor(C_NAVY)
        canvas.rect(0, 0, pw, 0.9*cm, fill=1, stroke=0)

        canvas.setFillColor(C_TEXT3)
        canvas.setFont("Helvetica", 7)
        canvas.drawString(1*cm, 0.3*cm,
                          f"{req.equipment_type}  |  {req.equipment_id}  |  "
                          f"{fmt_dmy(req.from_datetime)} → {fmt_dmy(req.to_datetime)}  |  "
                          f"Interval: {req.interval}  |  Function: {req.agg_function.upper()}")
        canvas.drawRightString(pw - 1*cm, 0.3*cm,
                               "GE Vernova Solar Monitoring — Confidential")

        canvas.restoreState()

    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.8*cm, bottomMargin=1.2*cm,
    )

    content = []

    # ── Details block ─────────────────────────────────────────────────────
    det_style = ParagraphStyle(
        "Det", fontSize=8.5, textColor=C_TEXT2,
        fontName="Helvetica", leading=15, spaceAfter=4,
    )
    details = (
        f"<b><font color='#00D4AA'>Equipment Type:</font></b> {req.equipment_type}"
        f"&nbsp;&nbsp;&nbsp;<b><font color='#00D4AA'>Equipment ID:</font></b> {req.equipment_id}"
        f"&nbsp;&nbsp;&nbsp;<b><font color='#00D4AA'>From:</font></b> {fmt_dmy(req.from_datetime)}"
        f"&nbsp;&nbsp;&nbsp;<b><font color='#00D4AA'>To:</font></b> {fmt_dmy(req.to_datetime)}<br/>"
        f"<b><font color='#00D4AA'>Interval:</font></b> {req.interval}"
        f"&nbsp;&nbsp;&nbsp;<b><font color='#00D4AA'>Function:</font></b> {req.agg_function.upper()}"
        f"&nbsp;&nbsp;&nbsp;<b><font color='#00D4AA'>Total Records:</font></b> {result.get('total_records', 0):,}"
    )
    content.append(Paragraph(details, det_style))
    content.append(Spacer(1, 0.2*cm))
    content.append(HRFlowable(width="100%", thickness=1,
                               color=C_ACCENT, spaceAfter=0.25*cm))

    # ── Table ─────────────────────────────────────────────────────────────
    headers    = [col_header(c) for c in cols]
    table_data = [headers]

    for row in rows:
        r = []
        for col in cols:
            val = row.get(col)
            if col == "timestamp":
                r.append(fmt_dmy(val) if val else "—")
            elif val is None:
                r.append("—")
            elif isinstance(val, float):
                r.append(f"{val:.3f}")
            elif isinstance(val, int):
                r.append(str(val))
            else:
                r.append(str(val))
        table_data.append(r)

    # Column widths
    usable_w = pw - 2*cm
    ts_w     = 3.8*cm
    n_cols   = len(cols)
    other_w  = (usable_w - ts_w) / max(n_cols - 1, 1) if n_cols > 1 else usable_w
    col_widths = [ts_w if c == "timestamp" else other_w for c in cols]

    n_rows   = len(table_data)
    t_style  = [
        # Header
        ("BACKGROUND",    (0, 0), (-1, 0),  C_NAVY),
        ("TEXTCOLOR",     (0, 0), (-1, 0),  C_ACCENT),
        ("FONTNAME",      (0, 0), (-1, 0),  "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, 0),  7),
        ("ALIGN",         (0, 0), (-1, 0),  "CENTER"),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LINEBELOW",     (0, 0), (-1, 0),  1.5, C_ACCENT),
        ("ROWHEIGHT",     (0, 0), (-1, 0),  20),
        # Data
        ("FONTNAME",      (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",      (0, 1), (-1, -1), 7),
        ("TEXTCOLOR",     (0, 1), (-1, -1), C_TEXT2),
        ("ALIGN",         (0, 1), (0,  -1), "CENTER"),
        ("ALIGN",         (1, 1), (-1, -1), "RIGHT"),
        ("ROWHEIGHT",     (0, 1), (-1, -1), 14),
        # Grid
        ("GRID",          (0, 0), (-1, -1), 0.25, C_BORDER),
        ("TOPPADDING",    (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
    ]

    for ri in range(1, n_rows):
        bg = C_EVEN if ri % 2 == 0 else C_ODD
        t_style.append(("BACKGROUND", (0, ri), (-1, ri), bg))

    t = Table(table_data, colWidths=col_widths, repeatRows=1, splitByRow=True)
    t.setStyle(TableStyle(t_style))
    content.append(t)

    doc.build(content, onFirstPage=on_page, onLaterPages=on_page)
    output.seek(0)
    filename = f"{req.equipment_id}_Report_{today_dmy()}.pdf"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Access-Control-Expose-Headers": "Content-Disposition",
        }
    )


# ── Summary ───────────────────────────────────────────────────────────────────
@router.post("/summary")
def get_summary(req: ReportDataRequest, db: Session = Depends(get_db)):
    from app.repositories.reports_repository import EQUIPMENT_REGISTRY, _safe_name
    from sqlalchemy import text

    config   = EQUIPMENT_REGISTRY.get(req.equipment_type, {})
    tag_meta = config.get("tags", {})

    column_names = []
    for tag in req.tags:
        if tag in tag_meta:
            column_names.append(tag)
        else:
            found = next(
                (col for col, m in tag_meta.items() if m["label"] == tag), tag
            )
            column_names.append(found)

    for col in column_names:
        if not _safe_name(col):
            raise HTTPException(400, detail=f"Invalid tag: {col}")

    table     = req.equipment_id
    agg_parts = []
    for col in column_names:
        agg_parts.extend([
            f"AVG(CAST([{col}] AS FLOAT)) AS [{col}_avg]",
            f"MIN(CAST([{col}] AS FLOAT)) AS [{col}_min]",
            f"MAX(CAST([{col}] AS FLOAT)) AS [{col}_max]",
            f"SUM(CAST([{col}] AS FLOAT)) AS [{col}_sum]",
        ])

    sql = text(f"""
        SELECT COUNT(*) AS record_count, {', '.join(agg_parts)}
        FROM [{table}]
        WHERE TimeCol BETWEEN :from_dt AND :to_dt
    """)
    try:
        row = db.execute(sql, {
            "from_dt": req.from_datetime.strftime("%Y-%m-%d %H:%M:%S"),
            "to_dt":   req.to_datetime.strftime("%Y-%m-%d %H:%M:%S"),
        }).fetchone()
        if not row:
            return {"record_count": 0, "stats": {}}
        stats = {"record_count": row[0]}
        idx   = 1
        for col in column_names:
            label = tag_meta.get(col, {}).get("label", col)
            stats[label] = {
                "avg": round(float(row[idx]),   3) if row[idx]   is not None else None,
                "min": round(float(row[idx+1]), 3) if row[idx+1] is not None else None,
                "max": round(float(row[idx+2]), 3) if row[idx+2] is not None else None,
                "sum": round(float(row[idx+3]), 3) if row[idx+3] is not None else None,
            }
            idx += 4
        return stats
    except Exception as e:
        logger.error(f"Summary error: {e}", exc_info=True)
        raise HTTPException(500, detail=str(e))